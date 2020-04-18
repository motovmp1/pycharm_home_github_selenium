# from class 24 and 25 for excel file use
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color

redFill = PatternFill(start_color='FFFF0000',
                      end_color='FFFF0000',
                      fill_type='solid')


def get_row_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.max_row


def get_column_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.max_row


def read_data_file(file, sheet_name, row_num, column_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.cell(row=row_num, column=column_num).value


def write_data_file(file, sheet_name, row_num, column_num, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    sheet.cell(row=row_num, column=column_num).value = data
    sheet['A1'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    # sheet['A1'].fill = redFill
    workbook.save(file)


def change_cell_color_fail(file, sheet_name, row_num, column_num, data, cell):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    sheet.cell(row=row_num, column=column_num).value = data
    sheet[cell].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    # sheet['A1'].fill = redFill
    workbook.save(file)


def change_cell_color_pass(file, sheet_name, row_num, column_num, data, cell):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    sheet.cell(row=row_num, column=column_num).value = data
    sheet[cell].fill = PatternFill(start_color="CCFFE5", end_color="CCFFE5", fill_type="solid")
    # sheet['A1'].fill = redFill
    workbook.save(file)