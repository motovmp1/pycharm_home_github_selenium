import openpyxl


book = openpyxl.load_workbook("/home/elsys/Desktop/Teste/pycharm_home_github_selenium/excel_selenium/pythondemo.xlsx")
# This is make active
sheet = book.active

cell = sheet.cell(row=1, column=2)

print(cell.value)

# write into cell
sheet.cell(row=2, column=2).value = "Vinicius"

cell2 = sheet.cell(row=2, column=2)
print(cell2)
print(cell2.value)

print(sheet.max_row)
print(sheet.max_column)

# value of cell
print(sheet['A5'])
# value inside the cell
print(sheet['A5'].value)

print("###########################################################")
# this is max number
for interaction in range(1, sheet.max_row+1):
    print(sheet.cell(row=interaction, column=1).value, end=" | ")
print("")
print("############################################################")




print("###########################################################")
# this is max number
for interaction in range(1, sheet.max_row+1):
    # quando o teste bate imprimi apenas a linha que interessa
    if sheet.cell(row=interaction, column=1).value == "Testcase2":
        for join in range(2, sheet.max_column+1):
            print(sheet.cell(row=interaction, column=join).value, end=" | ")
print("")
print("############################################################")


cell = sheet['A1': 'D7']

for c1, c2, c3, c4 in cell:
    print("{0:8}    {1:8}   {2:8}    {3:8}".format(c1.value, c2.value, c3.value, c4.value))

print("############################################################")